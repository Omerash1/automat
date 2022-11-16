from re import L
from types import CellType
import openpyxl as ox
from openpyxl import workbook
from openpyxl import load_workbook
import pyautogui
import time

#wb = workbook() 
#dest_filename = 'testwb.xlsx'
wb = load_workbook(filename=r"/Users/omerash/Downloads/automa-main/report3.xlsx")
ws1 = wb.active
lst=[]


#while CellType.value != None:

ws=wb.active
names=ws['X']

def gonext():
    l = 0
    for x in names:
        if x.value is None:
            for c in range(l + 1):
                names.pop(c.index())
            gonext()
            break
        l = l+1
        lst.append(x.value)
#pyautogui.moveTo(150, 100, duration=1)
#pyautogui.leftClick()
gonext()

#for x in lst:
    #pyautogui.write(x)
    #pyautogui.press("enter")

def automa():
    pyautogui.moveTo(150, 100, duration=1)
    pyautogui.leftClick()
    for x in lst:
        pyautogui.write(x)
        pyautogui.press("enter")
        if lst.index(x) == 29:
            for x in range(30):
                lst.pop(x)
            automa()
            break
automa()







